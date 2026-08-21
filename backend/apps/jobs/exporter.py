import io
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from apps.jobs.models import Job
from apps.skills.models import Skill
from apps.jobs.matcher import calculate_job_match

def generate_jobs_excel_workbook(user, queryset=None):
    """
    Generates a production-grade .xlsx Excel Workbook containing:
    - Sheet 1: Job Applications (No, Company, Role, Job Link, Applied Date, Status, Match Score, Required Skills, Missing Skills, My Skills)
    - Sheet 2: Skills Profile & Gap Analysis
    """
    if queryset is None:
        queryset = Job.objects.filter(user=user).prefetch_related('job_skills').order_by('-applied_date')

    wb = openpyxl.Workbook()

    # --- SHEET 1: JOB APPLICATIONS ---
    ws1 = wb.active
    ws1.title = "Job Applications"
    ws1.views.sheetView[0].showGridLines = True

    headers = [
        "No", "Company", "Role", "Job Link", "Applied Date",
        "Status", "Match Score", "Required Skills", "Missing Skills", "My Skills"
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws1.append(headers)
    ws1.row_dimensions[1].height = 28

    for col_num in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Fill Rows
    for index, job in enumerate(queryset, start=1):
        score, matching_skills, missing_skills, _ = calculate_job_match(user, job)
        required_skills_list = list(job.job_skills.values_list('skill_name', flat=True))

        required_str = ", ".join(required_skills_list)
        missing_str = ", ".join(missing_skills)
        matching_str = ", ".join(matching_skills)
        match_score_str = f"{score}%"

        row = [
            index,
            job.company,
            job.role,
            job.job_url or "",
            str(job.applied_date),
            job.get_status_display(),
            match_score_str,
            required_str,
            missing_str,
            matching_str
        ]
        ws1.append(row)

        current_row = index + 1
        ws1.row_dimensions[current_row].height = 22

        # Style data cells
        for col_num in range(1, len(headers) + 1):
            c = ws1.cell(row=current_row, column=col_num)
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)

            # Center justify No, Date, Status, Match Score
            if col_num in [1, 5, 6, 7]:
                c.alignment = Alignment(horizontal="center", vertical="center")

            # Green highlight for high match score (>=80%)
            if col_num == 7 and score >= 80:
                c.font = Font(bold=True, color="047857")

    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)


    # --- SHEET 2: SKILLS SUMMARY ---
    ws2 = wb.create_sheet(title="Skills Profile")
    ws2.views.sheetView[0].showGridLines = True
    ws2.append(["Skill Name", "Category", "Proficiency", "Source"])
    ws2.row_dimensions[1].height = 26

    for col_num in range(1, 5):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    skills_qs = Skill.objects.filter(user=user).order_by('category', 'name')
    for s_idx, skill in enumerate(skills_qs, start=2):
        ws2.append([skill.name, skill.get_category_display(), skill.get_proficiency_display(), skill.get_source_display()])
        ws2.row_dimensions[s_idx].height = 20

    for col in ws2.columns:
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = 22

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="JobOS_Applications_Export.xlsx"'
    return response


def generate_jobs_csv(user, queryset=None):
    """
    Generates a standard .csv export file.
    """
    if queryset is None:
        queryset = Job.objects.filter(user=user).prefetch_related('job_skills').order_by('-applied_date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="JobOS_Applications_Export.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "No", "Company", "Role", "Job Link", "Applied Date",
        "Status", "Match Score", "Required Skills", "Missing Skills", "My Skills"
    ])

    for index, job in enumerate(queryset, start=1):
        score, matching_skills, missing_skills, _ = calculate_job_match(user, job)
        required_skills_list = list(job.job_skills.values_list('skill_name', flat=True))

        writer.writerow([
            index,
            job.company,
            job.role,
            job.job_url or "",
            str(job.applied_date),
            job.get_status_display(),
            f"{score}%",
            ", ".join(required_skills_list),
            ", ".join(missing_skills),
            ", ".join(matching_skills)
        ])

    return response
